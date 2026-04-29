"""
Dosen Scraper - refactored from pddikti_scraping_update_copy.py
Accepts prodi_keywords as parameter instead of hardcoded constant.
"""

import requests
import time
import json
import os
import re
import sys
import threading
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import urllib.parse
try:
    from scraper.diktis_data import (
        is_diktis, refresh_ptkin_whitelist, classify_pt_from_name,
        classify_from_pembina, PTKIN_SET,
    )
except ImportError:
    from diktis_data import (
        is_diktis, refresh_ptkin_whitelist, classify_pt_from_name,
        classify_from_pembina, PTKIN_SET,
    )

BASE_URL = "https://api-pddikti.kemdiktisaintek.go.id"
HEADERS = {
    "Origin": "https://pddikti.kemdiktisaintek.go.id",
    "Referer": "https://pddikti.kemdiktisaintek.go.id/",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "application/json",
}
TIMEOUT = int(os.environ.get("PDDIKTI_TIMEOUT", 30))
MAX_RETRIES = max(1, int(os.environ.get("PDDIKTI_MAX_RETRIES", 3)))
RETRY_DELAY = int(os.environ.get("PDDIKTI_RETRY_DELAY", 2))
REQUEST_DELAY = float(os.environ.get("PDDIKTI_REQUEST_DELAY", 0.15))
MAX_WORKERS = int(os.environ.get("PDDIKTI_MAX_WORKERS", 5))

COLOR_DOSEN_HEADER = "1F4E79"
COLOR_DOSEN_ALT_ROW = "D6E4F0"
COLOR_DOSEN_TITLE = "2E75B6"

COLOR_PRODI_HEADER = "228B22"  # Forest Green
COLOR_PRODI_ALT_ROW = "E8F5E9" # Light Green
COLOR_PRODI_TITLE = "2E8B57"   # Sea Green


def normalize(name):
    """Normalize prodi name for comparison: uppercase, remove special chars."""
    name = name.upper()
    name = re.sub(r"[`'\u2018\u2019\u201A\u201B\u201C\u201D]", "", name)
    return " ".join(name.split())


def _clean_provinsi(raw: str) -> str:
    """Bersihkan nama provinsi dari API PDDikti.
    Contoh: 'Prov. D.K.I. Jakarta' → 'DKI JAKARTA'
            'Prov. Jawa Barat'     → 'JAWA BARAT'
    """
    if not raw:
        return ""
    s = str(raw).strip()
    # Hilangkan prefix "Prov." / "Prov"
    s = re.sub(r"^Prov\.?\s*", "", s, flags=re.IGNORECASE)
    # Hilangkan titik dalam singkatan (D.K.I. → DKI, D.I. → DI)
    s = s.replace(".", "")
    return " ".join(s.split()).upper()


def get_semesters(num_fallbacks=4):
    now = datetime.now()
    year, month = now.year, now.month
    if month == 1:
        current = f"{year-1}1"
    elif 2 <= month <= 8:
        current = f"{year-1}2"
    else:
        current = f"{year}1"
    sems, yr, t = [], int(current[:4]), int(current[4])
    for _ in range(num_fallbacks + 1):
        sems.append(f"{yr}{t}")
        if t == 2:
            t = 1
        else:
            t, yr = 2, yr - 1
    return sems[0], sems[1:]


def make_session():
    s = requests.Session()
    a = requests.adapters.HTTPAdapter(pool_connections=20, pool_maxsize=100, max_retries=3)
    s.mount("https://", a)
    s.mount("http://", a)
    return s


def fetch_api(session, endpoint, retries=MAX_RETRIES, stop_event=None):
    url = f"{BASE_URL}/{endpoint}"
    last_err = None
    for attempt in range(retries):
        try:
            r = session.get(url, headers=HEADERS, timeout=TIMEOUT)
            r.raise_for_status()
            data = r.json()
            if isinstance(data, dict) and data.get("message") == "Not Found":
                return None
            return data
        except requests.exceptions.Timeout as e:
            last_err = e
            if attempt < retries - 1:
                if stop_event: stop_event.wait(RETRY_DELAY * (attempt + 1))
                else: time.sleep(RETRY_DELAY * (attempt + 1))
        except requests.exceptions.RequestException as e:
            last_err = e
            if attempt < retries - 1:
                # Backoff lebih lama jika terkena Rate Limit 429
                if e.response is not None and e.response.status_code == 429:
                    if stop_event: stop_event.wait(5)
                    else: time.sleep(5)
                else:
                    if stop_event: stop_event.wait(RETRY_DELAY * (attempt + 1))
                    else: time.sleep(RETRY_DELAY * (attempt + 1))
        except json.JSONDecodeError:
            return None
            
    # Jika gagal total gara-gara server PDDikti (nge-lag/rate-limit)
    if last_err:
        raise Exception("Gagal terhubung ke server PDDikti (Server sibuk/rate-limit). Mohon jeda beberapa saat lalu coba lagi.")
    return None


PT_CACHE = {}
PT_CACHE_LOCK = threading.Lock()


def _fallback_pt_info(pt_query):
    """Helper: kembalikan klasifikasi fallback dari nama PT saja (tanpa API)."""
    info = classify_pt_from_name(pt_query)
    info.setdefault("pembina", "")
    info.setdefault("provinsi_pt", "")
    return info


def get_pt_info(session, pt_query, stop_event=None):
    """
    Mengidentifikasi: PTN/PTS, PTKIN/NON-PTKIN, DIKTI/DIKTIS.

    Strategi klasifikasi (2-tier):
      Primary : classify_from_pembina() — klasifikasi langsung dari field
                'pembina' API PDDikti (sumber paling otoritatif).
                PTA Islam Negeri/Swasta → DIKTIS, LLDIKTI/PTN → DIKTI.
      Fallback: Heuristik 3-lapis saat pembina kosong / tidak dikenal
                (whitelist PTKIN → keyword nama → keyword pembina/kelompok)

    Jika API PDDikti gagal di tahap mana pun (timeout / not found),
    fungsi tetap mengembalikan klasifikasi best-effort dari nama PT
    menggunakan classify_pt_from_name() — kolom Excel tidak akan kosong.
    """
    if not pt_query or not pt_query.strip():
        return {}

    # Thread-safe cache read
    with PT_CACHE_LOCK:
        if pt_query in PT_CACHE:
            return PT_CACHE[pt_query]

    def _cache_and_return(info):
        with PT_CACHE_LOCK:
            PT_CACHE[pt_query] = info
        return info

    # ── Step 1: Cari ID kampus di PDDikti ─────────────────────────
    quoted_pt = urllib.parse.quote(pt_query)
    results = fetch_api(session, f"pencarian/pt/{quoted_pt}", stop_event=stop_event)
    if not results:
        return _cache_and_return(_fallback_pt_info(pt_query))

    pt_id = results[0].get("id")
    if not pt_id:
        return _cache_and_return(_fallback_pt_info(pt_query))

    # ── Step 2: Ambil detail PT (pembina, kelompok, provinsi) ─────
    detail = fetch_api(session, f"pt/detail/{pt_id}", stop_event=stop_event)
    if not detail:
        return _cache_and_return(_fallback_pt_info(pt_query))

    pembina = (detail.get("pembina") or "").strip()
    kelompok = (detail.get("kelompok") or "").strip()
    provinsi_pt = _clean_provinsi(detail.get("provinsi_pt", ""))
    pt_upper = pt_query.upper().strip()

    # ── Step 3A (Primary): Klasifikasi langsung dari field pembina ──
    # Field pembina dari PDDikti API adalah sumber paling otoritatif:
    #   "PTA Islam Negeri"  → DIKTIS, PTN, PTKIN   (Kemenag)
    #   "PTA Islam Swasta"  → DIKTIS, PTS, NON PTKIN (Kemenag)
    #   "LLDIKTI [X]"       → DIKTI,  PTS, NON PTKIN (Kemendikbudristek)
    #   "PTN"               → DIKTI,  PTN, NON PTKIN (Kemendikbudristek)
    cls = classify_from_pembina(pembina)
    if cls:
        # Safety net: jika kampus ada di whitelist PTKIN tapi pembina-nya
        # tidak terduga (edge case saat data PDDikti belum di-update)
        if pt_upper in PTKIN_SET and cls["ptkin_non"] != "PTKIN":
            cls["ptkin_non"] = "PTKIN"
            cls["ptn_pts"] = "PTN"
            cls["dikti_diktis"] = "DIKTIS"
        info = {**cls, "pembina": pembina, "provinsi_pt": provinsi_pt}
        return _cache_and_return(info)

    # ── Step 3B (Fallback): Heuristik saat pembina kosong / tidak dikenal ──
    if pt_upper in PTKIN_SET:
        is_negeri = True
        is_ptkin = True
    else:
        is_negeri = ("NEGERI" in pembina.upper() or "NEGERI" in kelompok.upper())
        is_islam_check = (
            any(k in pembina.upper() for k in ["ISLAM", "AGAMA"])
            or any(k in kelompok.upper() for k in ["ISLAM", "AGAMA"])
        )
        is_ptkin = is_negeri and is_islam_check

    is_diktis_result = is_diktis(pt_query, pembina=pembina, kelompok=kelompok)

    info = {
        "ptn_pts": "PTN" if is_negeri else "PTS",
        "ptkin_non": "PTKIN" if is_ptkin else "NON PTKIN",
        "dikti_diktis": "DIKTIS" if is_diktis_result else "DIKTI",
        "pembina": pembina,
        "provinsi_pt": provinsi_pt,
    }
    return _cache_and_return(info)


def search_all_prodi(session, prodi_keywords, cb, stop_event=None):
    cb("=" * 60)
    cb(f"STEP 1: Mencari {len(prodi_keywords)} Keyword Program Studi...")
    cb("=" * 60)
    all_prodi, seen_ids, seen_logical = [], set(), set()
    norm_keywords = [normalize(k) for k in prodi_keywords]

    for keyword in prodi_keywords:
        if stop_event and stop_event.is_set():
            raise Exception("Scraping dihentikan oleh pengguna.")
        cb(f"\n🔍 Mencari: {keyword}")
        if stop_event: stop_event.wait(REQUEST_DELAY)
        else: time.sleep(REQUEST_DELAY)
        quoted_kw = urllib.parse.quote(keyword)
        results = fetch_api(session, f"pencarian/prodi/{quoted_kw}", stop_event=stop_event)
        if not results:
            cb(f"   ⚠️ Tidak ditemukan")
            continue
        count = 0
        for prodi in results:
            if stop_event and stop_event.is_set():
                break
            prodi_id = prodi.get("id", "")
            prodi_name = prodi.get("nama", "").strip().upper()
            if prodi_id in seen_ids:
                continue
            seen_ids.add(prodi_id)
            # Normalized matching to handle apostrophe variants
            prodi_norm = normalize(prodi_name)
            is_valid = any(kn in prodi_norm or prodi_norm == kn for kn in norm_keywords)
            if not is_valid:
                continue
            pt_name = prodi.get("pt", "")
            jenjang = prodi.get("jenjang", "")
            logical_key = f"{prodi_norm}|{jenjang}|{pt_name}".upper()
            if logical_key in seen_logical:
                continue
            if stop_event: stop_event.wait(REQUEST_DELAY)
            else: time.sleep(REQUEST_DELAY)
            detail = fetch_api(session, f"prodi/detail/{prodi_id}", stop_event=stop_event)

            # detail bisa None HANYA JIKA PDDikti mengembalikan 'Not Found' atau Error JSON Parse
            if detail is None:
                continue

            status = (detail.get("status") or "").strip()
            if status.upper() != "AKTIF":
                continue
            seen_logical.add(logical_key)
            pt_info = get_pt_info(session, pt_name, stop_event=stop_event)
            # Fallback: gunakan provinsi dari PT detail kalau provinsi prodi kosong
            # (detail sudah dijamin not None di atas, jadi akses langsung)
            provinsi_val = _clean_provinsi(detail.get("provinsi", "")) or pt_info.get("provinsi_pt", "")
            all_prodi.append({
                "id": prodi_id,
                "nama": prodi.get("nama", ""),
                "jenjang": jenjang,
                "pt": pt_name,
                "pt_singkat": prodi.get("pt_singkat", ""),
                "keterangan": status,
                "akreditasi": detail.get("akreditasi", ""),
                "provinsi": provinsi_val,
                "ptn_pts": pt_info.get("ptn_pts", ""),
                "ptkin_non": pt_info.get("ptkin_non", ""),
                "dikti_diktis": pt_info.get("dikti_diktis", ""),
                "pembina": pt_info.get("pembina", ""),
            })
            count += 1
        cb(f"   ✅ {count} prodi baru (total: {len(all_prodi)})")
    cb(f"\n📊 Total prodi ditemukan: {len(all_prodi)}")
    return all_prodi


def fetch_dosen_homebase(session, prodi_list, semester, fallbacks, cb, stop_event=None):
    cb("\n" + "=" * 60)
    cb("STEP 2: Mengambil daftar dosen homebase per prodi")
    cb("=" * 60)
    all_dosen, seen_nidn, failed = [], set(), []
    for i, prodi in enumerate(prodi_list, 1):
        if stop_event and stop_event.is_set():
            raise Exception("Scraping dihentikan oleh pengguna.")
        prodi_id = prodi["id"]
        label = f"{prodi['nama']} ({prodi['jenjang']}) - {prodi['pt']}"
        cb(f"\n[{i}/{len(prodi_list)}] {label}")
        if stop_event: stop_event.wait(REQUEST_DELAY)
        else: time.sleep(REQUEST_DELAY)
        dosen_list = fetch_api(session, f"dosen/homebase/{prodi_id}?semester={semester}", stop_event=stop_event)
        sem_used = semester
        if not dosen_list:
            for prev in fallbacks:
                if stop_event: stop_event.wait(REQUEST_DELAY)
                else: time.sleep(REQUEST_DELAY)
                dosen_list = fetch_api(session, f"dosen/homebase/{prodi_id}?semester={prev}", stop_event=stop_event)
                if dosen_list:
                    cb(f"   ℹ️ Menggunakan semester {prev}")
                    sem_used = prev
                    break
            if not dosen_list:
                cb(f"   ⚠️ Tidak ada data dosen")
                failed.append(label)
                prodi["semester_lapor"] = "Belum Lapor"
                continue
        prodi["semester_lapor"] = sem_used
        count_new = 0
        for d in dosen_list:
            if stop_event and stop_event.is_set():
                break
            nidn = d.get("nidn", "")
            nuptk = d.get("nuptk", "")
            nama = d.get("nama_dosen", "")
            key = nidn if nidn else (nuptk if nuptk else nama)
            if key in seen_nidn:
                continue
            seen_nidn.add(key)
            all_dosen.append({
                "nama_dosen": nama, "nidn": nidn, "nuptk": nuptk,
                "pendidikan": d.get("pendidikan", ""), "status_aktif": d.get("status_aktif", ""),
                "status_pegawai": d.get("status_pegawai", ""), "ikatan_kerja": d.get("ikatan_kerja", ""),
                "prodi_asal": prodi["nama"], "jenjang_prodi": prodi["jenjang"],
                "pt_asal": prodi["pt"], "semester_data": sem_used,
            })
            count_new += 1
        cb(f"   ✅ {len(dosen_list)} dosen, {count_new} baru (total unik: {len(all_dosen)})")
    if failed:
        cb(f"\n⚠️ {len(failed)} prodi tidak punya data dosen")
    cb(f"\n📊 Total dosen unik: {len(all_dosen)}")
    return all_dosen


def fetch_single_profile(session, dosen, stop_event=None):
    nidn = dosen.get("nidn", "")
    nuptk = dosen.get("nuptk", "")
    nama = dosen.get("nama_dosen", "")
    result = {
        "Nama": nama, "Perguruan Tinggi": dosen.get("pt_asal", ""),
        "Jabatan Fungsional": "", "Status Ikatan Kerja": dosen.get("ikatan_kerja", ""),
        "Jenis Kelamin": "", "Program Studi": dosen.get("prodi_asal", ""),
        "Jenjang": dosen.get("jenjang_prodi", ""), "Pendidikan Terakhir": dosen.get("pendidikan", ""),
        "Status Aktifitas": dosen.get("status_aktif", ""), "NIDN": nidn,
        "NUPTK": nuptk, "Status Kepegawaian": dosen.get("status_pegawai", ""),
        "Semester Data": dosen.get("semester_data", ""),
    }
    search_key = nidn if nidn else (nuptk if nuptk else nama)
    if not search_key:
        return result
    if stop_event: stop_event.wait(REQUEST_DELAY)
    else: time.sleep(REQUEST_DELAY)
    search_results = fetch_api(session, f"pencarian/dosen/{search_key}", stop_event=stop_event)
    if not search_results:
        return result
    search_id = None
    for sr in search_results:
        if nidn and sr.get("nidn", "") == nidn:
            search_id = sr.get("id", ""); break
        elif nuptk and sr.get("nuptk", "") == nuptk:
            search_id = sr.get("id", ""); break
        elif sr.get("nama", "").upper() == nama.upper():
            search_id = sr.get("id", ""); break
    if not search_id:
        return result
    if stop_event: stop_event.wait(REQUEST_DELAY)
    else: time.sleep(REQUEST_DELAY)
    profile = fetch_api(session, f"dosen/profile/{search_id}", stop_event=stop_event)
    if profile and isinstance(profile, dict):
        profile_nidn = profile.get("nidn_dosen", "") or profile.get("nidn", "")
        profile_nuptk = profile.get("nuptk", "")
        if nidn and profile_nidn and profile_nidn != nidn:
            return result
        if not nidn and nuptk and profile_nuptk and profile_nuptk != nuptk:
            return result
        profil_nama = profile.get("nama_dosen", "")
        if profil_nama and nama:
            t1 = set(w for w in nama.lower().split() if len(w) > 2)
            t2 = set(w for w in profil_nama.lower().split() if len(w) > 2)
            if t1 and t2 and t1.isdisjoint(t2):
                return result
        result["Nama"] = profile.get("nama_dosen", nama)
        result["Perguruan Tinggi"] = profile.get("nama_pt", result["Perguruan Tinggi"])
        result["Jabatan Fungsional"] = profile.get("jabatan_akademik", "")
        result["Status Ikatan Kerja"] = profile.get("status_ikatan_kerja", result["Status Ikatan Kerja"])
        result["Jenis Kelamin"] = profile.get("jenis_kelamin", "")
        result["Pendidikan Terakhir"] = profile.get("pendidikan_tertinggi", result["Pendidikan Terakhir"])
        result["Status Aktifitas"] = profile.get("status_aktivitas", result["Status Aktifitas"])
    return result


def fetch_all_profiles(session, dosen_list, cb, stop_event=None):
    cb("\n" + "=" * 60)
    cb(f"STEP 3: Mengambil profil detail {len(dosen_list)} dosen...")
    cb("=" * 60)
    all_profiles, completed, failed = [], 0, 0
    start = time.time()
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(fetch_single_profile, session, d, stop_event): i + 1 for i, d in enumerate(dosen_list)}
        for future in as_completed(futures):
            if stop_event and stop_event.is_set():
                # cancel_futures hanya tersedia sejak Python 3.9
                if sys.version_info >= (3, 9):
                    executor.shutdown(wait=False, cancel_futures=True)
                else:
                    executor.shutdown(wait=False)
                raise Exception("Scraping dihentikan oleh pengguna.")
            completed += 1
            try:
                all_profiles.append(future.result())
            except Exception:
                failed += 1
            if completed % 50 == 0 or completed == len(dosen_list):
                elapsed = time.time() - start
                rate = completed / elapsed if elapsed > 0 else 1
                eta = (len(dosen_list) - completed) / rate
                pct = completed * 100 // len(dosen_list)
                cb(f"📊 Progress: {completed}/{len(dosen_list)} ({pct}%) — ETA: ~{eta:.0f}s")
    cb(f"\n✅ {len(all_profiles)} profil berhasil, {failed} gagal")
    return all_profiles


def _apply_header_style(ws, row, max_col, header_color):
    hf = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    for col in range(1, max_col + 1):
        c = ws.cell(row=row, column=col)
        c.font, c.fill, c.alignment, c.border = hf, hfill, ha, border


def _apply_data_style(ws, start_row, end_row, max_col, alt_row_color):
    alt = PatternFill(start_color=alt_row_color, end_color=alt_row_color, fill_type="solid")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    for row in range(start_row, end_row + 1):
        for col in range(1, max_col + 1):
            c = ws.cell(row=row, column=col)
            c.border = border
            c.alignment = Alignment(vertical="center")
            if (row - start_row) % 2 == 1:
                c.fill = alt


def _auto_width(ws, max_col):
    for col in range(1, max_col + 1):
        mx = max((len(str(c.value or "")) for row in ws.iter_rows(min_col=col, max_col=col) for c in row), default=10)
        ws.column_dimensions[get_column_letter(col)].width = min(mx + 4, 50)


def export_to_excel(profiles, prodi_list, semester, output_dir, cb):
    cb("\n" + "=" * 60)
    cb("STEP 4: Membuat file Excel...")
    cb("=" * 60)
    wb = Workbook()
    wb.remove(wb.active)
    time_str = datetime.now().strftime("%d %B %Y, %H:%M WIB")

    # --- Sheet 1: Data Dosen ---
    ws = wb.create_sheet("Data Dosen")
    cols = ["No", "Nama", "Perguruan Tinggi", "Jabatan Fungsional", "Status Ikatan Kerja",
            "Jenis Kelamin", "Program Studi", "Pendidikan Terakhir", "Status Aktifitas",
            "NIDN", "NUPTK", "Status Kepegawaian", "Jenjang", "Semester Data"]
    mc = len(cols)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=mc)
    ws.cell(1, 1, "📊 DATA DOSEN PROGRAM STUDI PILIHAN").font = Font(name="Calibri", bold=True, color=COLOR_DOSEN_TITLE, size=14)
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=mc)
    ws.cell(2, 1, f"Sumber: PDDikti API — Diambil pada {time_str}").font = Font(name="Calibri", italic=True, color="666666", size=9)
    ws.cell(2, 1).alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=mc)
    ws.cell(3, 1, f"Total Dosen: {len(profiles)} — Semester: {semester}").font = Font(name="Calibri", bold=True, size=10)
    ws.cell(3, 1).alignment = Alignment(horizontal="center")
    for col, h in enumerate(cols, 1):
        ws.cell(5, col, h)
    _apply_header_style(ws, 5, mc, COLOR_DOSEN_HEADER)
    sorted_profiles = sorted(profiles, key=lambda x: (x.get("Program Studi", ""), x.get("Nama", "")))
    for i, p in enumerate(sorted_profiles, 1):
        row = 5 + i
        vals = [i, p.get("Nama"), p.get("Perguruan Tinggi"), p.get("Jabatan Fungsional"),
                p.get("Status Ikatan Kerja"), p.get("Jenis Kelamin"), p.get("Program Studi"),
                p.get("Pendidikan Terakhir"), p.get("Status Aktifitas"), p.get("NIDN"),
                p.get("NUPTK"), p.get("Status Kepegawaian"), p.get("Jenjang"), p.get("Semester Data")]
        for col, val in enumerate(vals, 1):
            ws.cell(row, col, val)
    _apply_data_style(ws, 6, 5 + len(sorted_profiles), mc, COLOR_DOSEN_ALT_ROW)
    _auto_width(ws, mc)
    ws.freeze_panes = "B6"

    # --- Sheet 2: Daftar Prodi ---
    ws2 = wb.create_sheet("Daftar Prodi")
    p_cols = ["No", "Nama Prodi", "Jenjang", "Perguruan Tinggi", "Jumlah Dosen",
              "Keterangan", "Akreditasi Program Studi", "PTN/PTS",
              "PTKIN/NON PTKIN", "DIKTI/DIKTIS", "Pembina", "Provinsi",
              "Semester Laporan Terakhir"]
    mc2 = len(p_cols)
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=mc2)
    ws2.cell(1, 1, "📋 DAFTAR PROGRAM STUDI PILIHAN").font = Font(name="Calibri", bold=True, color=COLOR_PRODI_TITLE, size=14)
    ws2.cell(1, 1).alignment = Alignment(horizontal="center")
    for col, h in enumerate(p_cols, 1):
        ws2.cell(3, col, h)
    _apply_header_style(ws2, 3, mc2, COLOR_PRODI_HEADER)
    prodi_count = {}
    for p in sorted_profiles:
        key = f"{p.get('Program Studi','').upper()}|{p.get('Perguruan Tinggi','').upper()}"
        prodi_count[key] = prodi_count.get(key, 0) + 1
    for i, prodi in enumerate(sorted(prodi_list, key=lambda x: (x["nama"], x["pt"])), 1):
        row = 3 + i
        key = f"{prodi['nama'].upper()}|{prodi['pt'].upper()}"
        count = prodi_count.get(key, 0)
        vals = [i, prodi["nama"], prodi["jenjang"], prodi["pt"], count,
                prodi.get("keterangan"), prodi.get("akreditasi"), prodi.get("ptn_pts"),
                prodi.get("ptkin_non"), prodi.get("dikti_diktis"),
                prodi.get("pembina"), prodi.get("provinsi"),
                prodi.get("semester_lapor", "Belum Lapor")]
        for col, val in enumerate(vals, 1):
            ws2.cell(row, col, val)
    _apply_data_style(ws2, 4, 3 + len(prodi_list), mc2, COLOR_PRODI_ALT_ROW)
    _auto_width(ws2, mc2)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"data_dosen_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)
    
    # Selesai
    cb(f"✅ File Excel disimpan: {filename}")
    return filename


def run_dosen_scraper(prodi_keywords, output_dir, callback, stop_event=None):
    """Entry point. prodi_keywords: list of prodi name strings."""
    # PT_CACHE bersifat additive (data PT tidak berubah antar sesi),
    # sehingga TIDAK perlu di-clear — ini juga mencegah corruption
    # jika ada dua job berjalan bersamaan (C2 fix).

    # ── Auto-refresh whitelist PTKIN dari SPAN-PTKIN (sekali per sesi) ──
    refresh_ptkin_whitelist(log_fn=callback)

    session = make_session()
    semester, fallbacks = get_semesters()
    callback(f"\n⚙️  Semester: {semester}")
    callback(f"⚙️  Keywords terpilih: {len(prodi_keywords)}")
    callback(f"⚙️  Max Workers: {MAX_WORKERS}\n")

    prodi_list = search_all_prodi(session, prodi_keywords, callback, stop_event)
    if not prodi_list:
        raise Exception("Tidak ada prodi ditemukan. Periksa koneksi internet atau keyword.")

    dosen_list = fetch_dosen_homebase(session, prodi_list, semester, fallbacks, callback, stop_event)
    if not dosen_list:
        raise Exception("Tidak ada dosen ditemukan.")

    profiles = fetch_all_profiles(session, dosen_list, callback, stop_event)
    filename = export_to_excel(profiles, prodi_list, semester, output_dir, callback)

    callback(f"\n🎉 SELESAI! Total dosen: {len(profiles)}, Total prodi: {len(prodi_list)}")
    return filename
