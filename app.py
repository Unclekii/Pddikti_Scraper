import os
import uuid
import queue
import threading
import json
import time
from flask import Flask, render_template, request, jsonify, Response, send_from_directory
from functools import wraps
from werkzeug.utils import secure_filename

app = Flask(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Active jobs: {job_id: {'queue': Queue, 'status': str, 'filename': str, 'event': Event, 'start_time': float}}
JOBS = {}
JOBS_LOCK = threading.Lock()

def cleanup_stale_jobs():
    """Background task to remove jobs older than 1 hour to prevent memory leaks."""
    while True:
        time.sleep(600)  # Run every 10 minutes
        now = time.time()
        with JOBS_LOCK:
            stale_ids = [jid for jid, info in JOBS.items() if now - info.get('start_time', 0) > 3600]
            for jid in stale_ids:
                event = JOBS[jid].get("event")
                if event:
                    event.set()
                del JOBS[jid]

_cleanup_started = False

@app.before_request
def _start_cleanup_once():
    """Lazy-start cleanup thread on first request — compatible with Gunicorn pre-fork workers."""
    global _cleanup_started
    if not _cleanup_started:
        _cleanup_started = True
        threading.Thread(target=cleanup_stale_jobs, daemon=True).start()


# ──────────────────────────────────────────────
# B4: OPTIONAL BASIC AUTH
# ──────────────────────────────────────────────
AUTH_KEY = os.environ.get("PDDIKTI_AUTH_KEY", "").strip()

def check_auth(f):
    """Decorator: enforce Basic Auth only if PDDIKTI_AUTH_KEY env var is set."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not AUTH_KEY:
            return f(*args, **kwargs)
        auth = request.authorization
        if not auth or auth.username != "admin" or auth.password != AUTH_KEY:
            return Response(
                "Akses ditolak. Masukkan kredensial yang benar.",
                401,
                {"WWW-Authenticate": 'Basic realm="PDDikti Dashboard"'},
            )
        return f(*args, **kwargs)
    return decorated


# ──────────────────────────────────────────────
# ROUTES
# ──────────────────────────────────────────────
@app.route("/")
@check_auth
def index():
    return render_template("index.html")


@app.route("/api/fetch-prodi")
@check_auth
def api_fetch_prodi():
    """Fetch all prodi from all bidang ilmu."""
    from scraper.fetch_prodi import fetch_all_prodi
    try:
        data = fetch_all_prodi()
        return jsonify({"success": True, "data": data, "total": len(data)})
    except Exception as e:
        app.logger.exception("fetch-prodi failed")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/run-scraper", methods=["POST"])
@check_auth
def api_run_scraper():
    """Start dosen scraping job with selected prodi keywords."""
    # Robust JSON parsing: dukung payload object {"prodi":[...]} maupun array langsung [...]
    body = request.get_json(silent=True)

    if isinstance(body, dict):
        prodi_keywords = body.get("prodi", [])
    elif isinstance(body, list):
        prodi_keywords = body
    else:
        prodi_keywords = []

    # Sanitasi agar hanya string non-empty yang diproses
    prodi_keywords = [str(x).strip() for x in prodi_keywords if str(x).strip()]

    if not prodi_keywords:
        return jsonify({"success": False, "error": "Tidak ada prodi yang dipilih"}), 400

    job_id = str(uuid.uuid4())
    q = queue.Queue()
    stop_event = threading.Event()
    with JOBS_LOCK:
        JOBS[job_id] = {
            "queue": q, 
            "status": "running", 
            "filename": None, 
            "event": stop_event, 
            "start_time": time.time()
        }

    def worker():
        from scraper.dosen_scraper import run_dosen_scraper
        try:
            def cb(msg):
                # B3: Structured progress messages
                if isinstance(msg, dict) and msg.get("__progress__"):
                    try:
                        q.put({"type": "progress", "step": msg["step"], "current": msg["current"],
                               "total": msg["total"], "label": msg["label"]}, block=False)
                    except queue.Full:
                        pass
                    return
                # Non-blocking put: tetap kirim log meski stop_event aktif
                # agar log terakhir (cancellation) tidak hilang
                try:
                    q.put({"type": "log", "message": str(msg)}, block=False)
                except queue.Full:
                    pass  # Queue penuh — skip daripada deadlock

            filename = run_dosen_scraper(prodi_keywords, OUTPUT_DIR, cb, stop_event)
            with JOBS_LOCK:
                if job_id in JOBS:
                    JOBS[job_id]["status"] = "done"
                    JOBS[job_id]["filename"] = filename
            q.put({"type": "done", "filename": filename})
        except Exception as e:
            with JOBS_LOCK:
                if job_id in JOBS:
                    JOBS[job_id]["status"] = "error"
            q.put({"type": "error", "message": str(e)})
        finally:
            q.put(None)  # sentinel

    t = threading.Thread(target=worker, daemon=True)
    t.start()
    return jsonify({"success": True, "job_id": job_id})


@app.route("/api/stop-scraper/<job_id>", methods=["POST"])
@check_auth
def api_stop_scraper(job_id):
    """Stop an active scraping job."""
    with JOBS_LOCK:
        if job_id in JOBS and JOBS[job_id].get("event"):
            JOBS[job_id]["event"].set()
            JOBS[job_id]["status"] = "cancelled"
            return jsonify({"success": True, "message": "Scraping dibatalkan"})
    return jsonify({"success": False, "error": "Job tidak ditemukan atau sudah selesai"}), 404


@app.route("/api/stream/<job_id>")
def api_stream(job_id):
    """SSE endpoint to stream scraping logs."""
    with JOBS_LOCK:
        if job_id not in JOBS:
            return jsonify({"error": "Job not found"}), 404
        q = JOBS[job_id]["queue"]

    def generate():
        try:
            yield "data: {\"type\": \"connected\"}\n\n"
            while True:
                try:
                    # Heartbeat setiap 30s untuk menghindari Nginx/proxy timeout default 60s
                    msg = q.get(timeout=30)
                    if msg is None:
                        break
                    yield f"data: {json.dumps(msg)}\n\n"
                except queue.Empty:
                    yield "data: {\"type\": \"heartbeat\"}\n\n"
        finally:
            # DO NOT DELETE FROM JOBS! If the SSE stream disconnects, 
            # the background worker should continue running.
            pass

    return Response(
        generate(),
        content_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.route("/api/outputs")
@check_auth
def api_outputs():
    """List output Excel files."""
    files = []
    if not os.path.exists(OUTPUT_DIR):
        return jsonify([])
    for f in os.listdir(OUTPUT_DIR):
        # Filter temp/lock files Excel (diawali ~$) & hanya ambil .xlsx valid
        if f.endswith(".xlsx") and not f.startswith("~$") and not f.startswith("."):
            fp = os.path.join(OUTPUT_DIR, f)
            try:
                files.append({
                    "name": f,
                    "size": os.path.getsize(fp),
                    "modified": os.path.getmtime(fp)
                })
            except OSError:
                # Skip file yang tiba-tiba hilang / permission error
                continue
    files.sort(key=lambda x: x["modified"], reverse=True)
    return jsonify(files)


# B5: Download endpoint hardening — validasi konsisten dengan delete endpoint
@app.route("/api/download/<filename>")
@check_auth
def api_download(filename):
    """Download an output Excel file."""
    filename = secure_filename(filename)
    if not filename.lower().endswith(".xlsx"):
        return jsonify({"success": False, "error": "Only .xlsx files can be downloaded"}), 400
    return send_from_directory(OUTPUT_DIR, filename, as_attachment=True)


@app.route("/api/delete-file/<filename>", methods=["DELETE"])
@check_auth
def api_delete_file(filename):
    """Delete an output Excel file (.xlsx only)."""
    filename = secure_filename(filename)
    if not filename.lower().endswith(".xlsx"):
        return jsonify({"success": False, "error": "Only .xlsx files can be deleted"}), 400
    try:
        fp = os.path.join(OUTPUT_DIR, filename)
        if os.path.exists(fp):
            os.remove(fp)
            return jsonify({"success": True})
        return jsonify({"success": False, "error": "File not found"}), 404
    except Exception as e:
        app.logger.exception("delete-file failed")
        return jsonify({"success": False, "error": str(e)}), 500


# ──────────────────────────────────────────────
# ANALYTICS API — Parse Excel & return JSON
# ──────────────────────────────────────────────
@app.route("/api/analyze/<filename>")
@check_auth
def api_analyze(filename):
    """Parse an output Excel file and return both sheets as JSON for analytics."""
    filename = secure_filename(filename)
    if not filename.lower().endswith(".xlsx"):
        return jsonify({"success": False, "error": "Only .xlsx files"}), 400

    fp = os.path.join(OUTPUT_DIR, filename)
    if not os.path.exists(fp):
        return jsonify({"success": False, "error": "File not found"}), 404

    try:
        from openpyxl import load_workbook
        wb = load_workbook(fp, read_only=True, data_only=True)

        result = {"dosen": [], "prodi": [], "metadata": {"filename": filename}}

        # --- Parse Sheet "Data Dosen" ---
        if "Data Dosen" in wb.sheetnames:
            ws = wb["Data Dosen"]
            # Header di baris 5
            headers = []
            for cell in ws[5]:
                headers.append(cell.value)
            for row in ws.iter_rows(min_row=6, values_only=True):
                if not row or row[0] is None:
                    continue
                record = {}
                for i, val in enumerate(row):
                    if i < len(headers) and headers[i]:
                        record[headers[i]] = val if val is not None else ""
                result["dosen"].append(record)

        # --- Parse Sheet "Daftar Prodi" ---
        if "Daftar Prodi" in wb.sheetnames:
            ws2 = wb["Daftar Prodi"]
            # Header di baris 3
            headers2 = []
            for cell in ws2[3]:
                headers2.append(cell.value)
            for row in ws2.iter_rows(min_row=4, values_only=True):
                if not row or row[0] is None:
                    continue
                record = {}
                for i, val in enumerate(row):
                    if i < len(headers2) and headers2[i]:
                        record[headers2[i]] = val if val is not None else ""
                result["prodi"].append(record)

        wb.close()
        result["metadata"]["total_dosen"] = len(result["dosen"])
        result["metadata"]["total_prodi"] = len(result["prodi"])
        return jsonify({"success": True, **result})

    except Exception as e:
        app.logger.exception("analyze failed")
        return jsonify({"success": False, "error": str(e)}), 500


if __name__ == "__main__":
    print("=" * 50)
    print("  PDDikti Dashboard")
    print("  Buka browser: http://localhost:5000")
    print("=" * 50)
    app.run(debug=False, port=5000, threaded=True)
